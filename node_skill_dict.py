import pandas as pd
import numpy as np
import openpyxl
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import warnings, os, json

warnings.filterwarnings("ignore")
OUT_DIR = r"C:\Users\karet\3_jobs"

# ── 1. Matrix nodes ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(r"c:\Users\karet\Downloads\Skill_Matrix.xlsx")
ws = wb.active
nodes = []
cat = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value: cat = row[0].value
    if row[1].value: nodes.append({"category": cat, "type": "hard", "node": row[1].value.strip()})
    if row[2].value: nodes.append({"category": cat, "type": "soft", "node": row[2].value.strip()})
matrix = pd.DataFrame(nodes)

# ── 2. Collect ALL unique standardized skills from both years ────────────────
print("Loading 2023 skills...")
j23 = pd.read_csv(r"c:\Users\karet\Downloads\skill_drift_bundle\inputs\business_only_jobs_2023.csv")
s23 = set()
for skills in j23["standardized_skills"].dropna():
    for s in skills.split(","):
        s = s.strip()
        if s: s23.add(s)

print("Loading 2026 skills...")
j26 = pd.read_csv(r"c:\Users\karet\Downloads\skill_drift_bundle\inputs\business_only_jobs_2026.csv")
s26 = set()
for skills in j26["standardized_skills"].dropna():
    for s in skills.split(","):
        s = s.strip()
        if s: s26.add(s)

all_skills = sorted(s23 | s26)
print(f"Unique skills: {len(s23)} (2023) + {len(s26)} (2026) = {len(all_skills)} combined\n")

# ── 3. Embed nodes and skills ───────────────────────────────────────────────
print("Loading model...")
model = SentenceTransformer("all-MiniLM-L6-v2")

print(f"Encoding {len(matrix)} nodes...")
node_emb = model.encode(matrix["node"].tolist(), normalize_embeddings=True)

print(f"Encoding {len(all_skills)} skills...")
skill_emb = model.encode(all_skills, batch_size=256, show_progress_bar=True, normalize_embeddings=True)

# ── 4. For each skill, find best node match ─────────────────────────────────
sim = skill_emb @ node_emb.T
best_node_idx = np.argmax(sim, axis=1)
best_score = sim[np.arange(len(sim)), best_node_idx]

THRESHOLD = 0.45

# Build dictionary: node -> list of skills
node_to_skills = {row["node"]: [] for _, row in matrix.iterrows()}
skill_to_node = {}

mapping_rows = []
for i, skill in enumerate(all_skills):
    if best_score[i] >= THRESHOLD:
        n_idx = best_node_idx[i]
        node_name = matrix.iloc[n_idx]["node"]
        node_to_skills[node_name].append(skill)
        skill_to_node[skill] = node_name
        mapping_rows.append({
            "standardized_skill": skill,
            "matched_node": node_name,
            "node_category": matrix.iloc[n_idx]["category"],
            "node_type": matrix.iloc[n_idx]["type"],
            "similarity": round(float(best_score[i]), 3),
        })

mapping_df = pd.DataFrame(mapping_rows)
mapping_df.to_csv(os.path.join(OUT_DIR, "node_skill_dictionary.csv"), index=False)

# ── 5. Print dictionary ────────────────────────────────────────────────────
matched = len(mapping_df)
total = len(all_skills)
print(f"\n{'='*100}")
print(f"NODE -> SKILLS DICTIONARY  (threshold={THRESHOLD})")
print(f"Matched {matched}/{total} skills ({matched/total*100:.1f}%)")
print(f"{'='*100}\n")

for _, nrow in matrix.iterrows():
    skills = node_to_skills[nrow["node"]]
    if skills:
        print(f"[{nrow['category']:<10} {nrow['type']:<5}] {nrow['node']}")
        print(f"  ({len(skills)} skills): {', '.join(skills[:10])}")
        if len(skills) > 10:
            print(f"  ... and {len(skills)-10} more")
        print()

# Also save as JSON for easy reuse (skills sorted alphabetically per node)
json_dict = {}
for _, nrow in matrix.iterrows():
    key = nrow["node"]
    json_dict[key] = {
        "category": nrow["category"],
        "type": nrow["type"],
        "skills": sorted(node_to_skills[key]),
    }
with open(os.path.join(OUT_DIR, "node_skill_dictionary.json"), "w", encoding="utf-8") as f:
    json.dump(json_dict, f, indent=2, ensure_ascii=False)

print(f"\nSaved: node_skill_dictionary.csv, node_skill_dictionary.json")
